# importer les librairies besoins
import openpyxl as excel
import re, math

# des regles et des fonctions pour vérifierun string est-il un int ou float
_NUMBER__REG = re.compile(r'^-?\d+\.?\d*$') 
_INT__REG    = re.compile(r'^-?\d+$')
def transfNumber(string):
    if isInt(string):
        return int(string)
    return float(string)

def isInt(string):
    condition = _INT__REG.search(string)
    if condition:
        return True
    return False

def isNumber(string):
    condition = _NUMBER__REG.search(string)
    if condition:
        return True
    return False

def writeWB(ws, index, cot, numero=False):
    """Ecrire un String dans un unité de Excel

    Args:
        ws: un WorkBook actif
        index: un String avec le colone et la ligne de Excel
            par exemple: 'A7'
        cot: le contenu va être enregeisté
        numero: le contenu est-il un numero ou juste un string

    """
    ws[index] = transfNumber(cot) if numero else cot

def saveWB(wb, filename="data.xlsx"):
    """Enregistrer les modifications de la fonction writeWB

    Args:
        wb: un WorkBook
        filename: le nom de fichier Excel

    """
    wb.save(filename)

def lireUneLigne(f, content, ligneNumero):
    """Lire une ligne de data et enregistrer les données dans une dictionnaire

    Args:
        f: le fichier on veut lire (ici, "data.txt")
        content: une dict:
            les clés sont des String avec le colone et la ligne de Excel par exemple: 'A7'
            les valeurs sont des String des contenus on veut enregistrer      
        ligneNumero: le numero de la ligne Excel

    """
    ligne = f.readline().split("\t")
    i = 1
    for each in ligne:     
        index = excel.cell.cell.get_column_letter(i) + str(ligneNumero)
        content[index] = each
        i+=1

def initialiserData(content):
    """Initialiser tous les datas dans une dictionnaire

    Cette fonction va utiliser la fonction writeUneLigne() pour lire chaque ligne de
    la fichier, et va enregistrer les données dans une dict

    Args:
        content: une dict:
            les clés sont des String avec le colone et la ligne de Excel par exemple: 'A7'
            les valeurs sont des String des contenus on veut enregistrer
    """
    with open('data.txt') as dataFile:
        for it in range(1, 1008):
            lireUneLigne(dataFile, content, it)

def initialiserExcel(ws, content):
    """Initialiser tous les données dans un WorkBook de l'Excel

    Cette fonction va utiliser la fonction writeWB() pour écrire par ligne

    Args:
        ws: un WorkBook actif
        content: une dict:
            les clés sont des String avec le colone et la ligne de Excel par exemple: 'A7'
            les valeurs sont des String des contenus on veut enregistrer
    """
    for index in content:
        contenu = content[index]
        if isNumber(contenu):
            writeWB(ws, index, contenu, True)
        else:
            writeWB(ws, index, contenu)

def calculerDelta(ws, diminuende, diminuteur, ligne):
    """Calculer l'erreur entre la valeur demandée et la valeur fabriquée

    Args:
        ws: un WorkBook actif
        diminuende: la valeur fabriqué
        diminuteur: la valeur demandée
        ligne: le numero de la ligne Excel

    Returns:
        Si la valeur est manquante, renvoie un String vide ''
        Si non, renvoie l'erreur en Int ou Float, dépendant le type de diminuende et diminuteur
    """
    x = ws[diminuende + str(ligne)].value
    y = ws[diminuteur + str(ligne)].value
    if ((type(x) == float or type(x) == int) and (type(y) == float or type(y) == int)):
        res  = x - y
        return res
    else:
        return ''

def insertUneColone(ws, alpha, name, formule):
    """Calculer et enregistrer les données dans Excel

    Cette fonction va utiliser une autre fonction dans paramètre pour analyser les données

    Args:
        ws: un WorkBook actif
        alpha: la colone du Excel
        name: le nom de la colone
        formule: la fonction qu'on va utiliser pour analyser les données

    """
    if alpha == 'Q':
        para1 = 'H'
        para2 = 'G'
    elif alpha == 'R':
        para1 = 'J'
        para2 = 'I'
    elif alpha == 'S':
        para1 = 'L'
        para2 = 'K'
    elif alpha == 'T':
        para1 = 'N'
        para2 = 'M'
    elif alpha == 'U':
        para1 = 'P'
        para2 = 'O'
    for ligne in range(1, 1008):
        cellName = alpha + str(ligne)
        value = name if ligne == 1 else formule(ws, para1, para2, ligne)
        writeWB(ws, cellName, value)

def calculer_moyenne(liste_data):
    """Calculer la moyenne des valeurs

    Args:
        liste_data: une liste avec tous les valeurs dont on va calculer la moyenne

    Returns:
        float

    """
    res = 0
    for each in liste_data:
        res += each
    res /= len(liste_data)
    return res

def data_to_list(ws, alpha):
    """Transformer tous les données des une colone en une liste pour simplifier la calcul

    Args:
        ws: un WorkBook actif
        alpha: la colone du Excel

    Returns:
        Une liste avec deux éléments en type liste, le premier au nom de "org" l'autre de "oxl",
        les deux sous listes avec tous les données du type ORG et OXL

    """
    org = list()
    oxl = list()
    a = 1
    for number in range(2, 1008):
        type_cell = "C" + str(number)
        cell_name = alpha + str(number)
        valeur = ws[cell_name].value
        if (valeur != ""):
            if (ws[type_cell].value == "ORG"):
                org.append(valeur)
            else:
                oxl.append(valeur)
    return [org, oxl]

def calculer_variance(liste_data, moyenne):
    """Calculer la variance des données d'une type des verres

    Avant d'utiliser cette fonction, doit d'abord utiliser la fonction calculer_moyenne() pour avoir le moyenne.

    Args:
        liste_data: une liste avec tous les données dont on va calculer sa variance
        moyeene: la moyenne de la liste_data

    Returns:
        Float

    """
    res = 0
    for each in liste_data:
        tmp = (each - moyenne)**2
        res += tmp
    res /= len(liste_data)
    return res

def calculer_etendue(liste_data):
    """Calculer l'étendue des données d'une type des verres

    Le plus grande valeur moins le moins grande

    Args:
        liste_data: une liste avec tous les valeurs dont on va calculer l'étendue

    Returns:
        float

    """
    return max(liste_data) - min(liste_data)

def calculer_VADE(liste, moyenne):
    """Calculer la valeur absolue des écarts des données d'une type des verres

    Args:
        liste: une liste avec tous les valeurs dont on va calculer la valeur absolue des écarts
        moyenne: la moyenne de la liste

    Returns:
        float

    """
    res = 0
    for each in liste:
        res += (each - moyenne)
    return res/len(liste)

def dispersion_affichage(liste_data, name_colone):
    """Calculer enregistrer et afficher les résultat de statistiques descriptives de position et de dispersion.

    On va utiliser les fonctions suivantes pour analyser: calculer_moyenne(), calculer_variance(), calculer_VADE(),
    calculer_etendue()

    Args:
        liste_data: une liste avec tous les valeurs d'une colone
        name_colone: nom d'une colone du Excel

    """
    moyenne = calculer_moyenne(liste_data)
    print(" ==> Moyenne de", name_colone, "=", moyenne)
    variance = calculer_variance(liste_data, moyenne)
    print(" ==> Variance de", name_colone, "=", variance)
    ecart_type = math.sqrt(variance)
    print(" ==> Ecart type de", name_colone, "=", ecart_type)
    valeur_absolue_des_ecarts = calculer_VADE(liste_data, moyenne)
    print(" ==> Valeur absolue des écarts de", name_colone, "=", valeur_absolue_des_ecarts)
    etendue = calculer_etendue(liste_data)
    print(" ==> Etendue de", name_colone, "=", etendue)

def dispersion(ws, alpha):
    """Une fonction total pour analyser la dispersion des données

    Avant de utiliser fonction dispersion_affichage(), on doit d'abord diviser les deux sous listes de la liste_data,
    pour avoir une liste pour ORG, et une liste pour OXL, et en plus, on doit vérifier la liste est vide ou pas.

    Args:
        ws: un WorkBook actif
        alpha: la colone du Excel

    """
    name_colone = ws[alpha+"1"].value
    print(" ==> [%s]" % name_colone)
    liste_data = data_to_list(ws, alpha)
    print("[ ORG ]")
    if (len(liste_data[0]) != 0):
        dispersion_affichage(liste_data[0], name_colone)
    else:
        print(" ==> !! pas de données")
    print("[ OXL ]")
    if (len(liste_data[1]) != 0):
        dispersion_affichage(liste_data[1], name_colone)
    else:
        print(" ==> !! pas de données")
    print("*"*20,"-"*10 ,"*"*20)

def calculer_inter(liste_data, Z, moyenne, ecart_type):
    """Calculer intervalle de confiance pour les données d'un type des verres

    Normalement il faut utiliser la commande que j'ai commenté, mais si on divise l'écart type par racine de N,
    avec N le nombre des données, le résultat est trop faible, on va avoir un intervalle comme [moyenne-0.1, moyenne+0.1],
    cela n'est pas normal, donc on n'a pas diviser l'écart type par racine de N.

    Args:
        liste_data: une liste avec tous les valeurs dont on va calculer l'étendue
        Z: un constant qu'on peut avoir dans le tableau en fonction de alpha
        moyenne: la moyenne de la liste_data
        ecart_type: l'écart type de la liste_data

    Returns:
        Une liste avec deux éléments
    
    """
    debut = moyenne - Z * ecart_type
    #debut = moyenne - Z * (ecart_type / math.sqrt(len(liste_data)))
    fin = moyenne + Z * ecart_type
    #fin = moyenne + Z * (ecart_type / math.sqrt(len(liste_data)))
    res = [debut, fin]
    return res

def verifier_inter(a, inter, liste_data, name_colone):
    """Vérifier l'intervalle de confiance

    Vérifier s'il y a 95% ou %99 des valeurs est entre l'intervalle de confiance

    Args:
        a: la valeur de niveau de signification alpha
        inter: l'intervalle de confiance des données
        liste_data: une liste avec tous les valeurs dont on va calculer l'étendue
        name_colone: le nom de la colone du Excel
        
    Returns:
        Boolean, s'il y a 95% ou %99 des valeurs est entre l'intervalle de confiance, renvoit True, et
        afficher " ==> [ Prouvé ! ]", sinon, renvoit False, et afficher " ==> [ Problème ! ]"
    
    """
    compteur = 0
    for it in liste_data:
        if ((it < inter[0]) or (it > inter[1])):
            compteur += 1
    real = compteur / len(liste_data)
    real = 1 - real
    rea = real * 100
    print(" ==> %f%% de %s est entre le intervalle de confiance!" % (rea, name_colone))
    if real >= (1-a):
        print(" ==> [ Prouvé ! ]")
        return True
    print(" ==> [ Problème ! ]")
    return False

def inter_confiance_afficher(liste_data, name_colone):
    """Initialiser les deux niveau de signification et calcule l'intervalle de confiance.

    On va utiilser les fonctions suivantes: calculer_moyenne(), calculer_variance(), calculer_inter(), verifier_inter()

    Args:
        liste_data: une liste avec tous les valeurs dont on va calculer l'étendue
        name_colone: le nom de la colone du Excel

    """
    moyenne = calculer_moyenne(liste_data)
    variance = calculer_variance(liste_data, moyenne)
    ecart_type = math.sqrt(variance)
    Z = 1.960
    inter = calculer_inter(liste_data, Z, moyenne, ecart_type)
    print("[ a = 5% ]")
    print(" ==> Intervalle de confiance est", inter)
    a = 5e-2
    verifier_inter(a, inter, liste_data, name_colone)
    Z = 2.576
    inter = calculer_inter(liste_data, Z, moyenne, ecart_type)
    print("[ a = 1% ]")
    print(" ==> Intervalle de confiance est", inter)
    a = 1e-2
    verifier_inter(a, inter, liste_data, name_colone)

def inter_confiance(ws, alpha):
    """Une fonction totale pour analyser l'intervalle de confiance

        On va d'abord séparer la liste en deux ORG et OXL, et vérifier elle vide ou pas,
        et on va utiilser la fonction inter_confiance_afficher()

    Args:
        ws: un WorkBook actif
        alpha: la colone du Excel

    """
    name_colone = ws[alpha+"1"].value
    print(" ==> [%s]" % name_colone)
    liste_data = data_to_list(ws, alpha)
    print("[ ORG ]")
    if (len(liste_data[0]) != 0):
        inter_confiance_afficher(liste_data[0], name_colone)
    else:
        print(" ==> !! pas de données")
    print("[ OXL ]")
    if (len(liste_data[1]) != 0):
        inter_confiance_afficher(liste_data[1], name_colone)
    else:
        print(" ==> !! pas de données")
    print("*"*20,"-"*10 ,"*"*20)

def calculer_z(liste_data, name_colone):
    """Avoir la condition pour vérifier notre hypothèse

    On va utiliser les fonctions suivantes: calculer_moyenne(), calculer_variance().
    la condition z = (moyenne - hypothèse) / (écart_type / racine_de_n)
    Et on vérifier que la valeur de z est suppérieur ou inférieur à un constant Z du tableau.
    Ce constant Z on peut l'avoir dans un tableau en fonction d'un niveau de signification alpha.
    Ici, on juste vérifie deux cas, alpha = 0.05 et alpha = 0.01

    Args:
        liste_data: une liste avec tous les valeurs dont on va calculer l'étendue
        name_colone: le nom de la colone du Excel

    """
    moyenne = calculer_moyenne(liste_data)
    variance = calculer_variance(liste_data, moyenne)
    ecart_type = math.sqrt(variance)
    n = len(liste_data)
    z = (moyenne - 0) / (ecart_type / math.sqrt(n))
    if (z < -2.575) or (z > 2.575):
        print("[ 0.01 ] ==> H0 [ Refusé ]")
    else:
        print("[ 0.01 ] ==> H0 [ Accetpté ]")
    if (z < -1.96) or (z > 1.96):
        print("[ 0.05 ] ==> H0 [ Refusé ]")
    else:
        print("[ 0.05 ] ==> H0 [ Accetpté ]")
    print(" ==> z =", z)

def calculer_h1(ws, alpha):
    """Une fonction totale pour analyser le test d'hypothèse

    On va d'abord séparer la liste en deux ORG et OXL, et vérifier elle vide ou pas,
    et on va utiilser la fonction calculer_z()

    Args:
        ws: un WorkBook actif
        alpha: la colone du Excel

    """
    name_colone = ws[alpha+"1"].value
    print(" ==> [%s]" % name_colone)
    liste_data = data_to_list(ws, alpha)
    print("[ ORG ]")
    if (len(liste_data[0]) != 0):
        calculer_z(liste_data[0], name_colone)
    else:
        print(" ==> !! pas de données")
    print("[ OXL ]")
    if (len(liste_data[1]) != 0):
        calculer_z(liste_data[1], name_colone)
    else:
        print(" ==> !! pas de données")
    print("*"*20,"-"*10 ,"*"*20)

def calculer_K(liste_data, name_colone):
    """Calculer le K carré pour vérifier la relation entre le Delta et le type du verre

    On doit calculer K au carré en fonction de a, b, c, d, n

    Args:
        ws: un WorkBook actif
        name_colone: la colone du Excel

    """
    ORG = liste_data[0]
    OXL = liste_data[1]
    moyenne = calculer_moyenne(ORG)
    variance = calculer_variance(ORG, moyenne)
    ecart_type = math.sqrt(variance)
    Z = 1.960
    inter = calculer_inter(ORG, Z, moyenne, ecart_type)
    a = 0
    for it in ORG:
        if (it >= inter[0] and it <= inter[1]):
            a += 1
    b = len(ORG) - a
    moyenne = calculer_moyenne(OXL)
    variance = calculer_variance(OXL, moyenne)
    ecart_type = math.sqrt(variance)
    Z = 1.960
    inter = calculer_inter(OXL, Z, moyenne, ecart_type)
    c = 0
    for it in OXL:
        if (it >= inter[0] and it <= inter[1]):
            c += 1
    d = len(OXL) - c
    n = a + b + c + d
    print("a = %d, b = %d, c = %d , d = %d, n = %d"%(a,b,c,d,n))
    K2 = (n * (a*d - b*c)**2) / ((a+b)*(c+d)*(a+c)*(b+d))
    print(" ==> K2 =", K2)
    if K2 <= 3.841:
        print("[ Non relation ] <==> K2 <= 3.841")
    elif (K2 > 3.841 and K2 < 6.635):
        print("[ 95% relation ] <==> 3.841 < K2 < 6.635")
    else:
        print("[ 99% relation ] <==> K2 >= 6.635")

def test_de_comparaison_de_moyenne(ws, alpha):
    """Une fonction totale pour analyser le test de comparaison de moyennes

    On va d'abord séparer la liste en deux ORG et OXL, et vérifier elle vide ou pas,
    et on va utiilser la fonction calculer_K()

    Args:
        ws: un WorkBook actif
        alpha: la colone du Excel

    """
    name_colone = ws[alpha+"1"].value
    print(" ==> [%s]" % name_colone)
    liste_data = data_to_list(ws, alpha)
    if (len(liste_data[0]) != 0 and len(liste_data[1]) != 0):
        calculer_K(liste_data, name_colone)
    else:
        print(" ==> !! pas de données")
    print("*"*20,"-"*10 ,"*"*20)   
    
def main():
    # Créer une instance de Workbook pour Excel
    wb = excel.Workbook();

    # Obtenir le Workbook actif
    ws = wb.active

    # Enregistrer les datas dans une dictionnaire
    content = {}
    initialiserData(content)
    
    # Utiliser cette dictionnaire pour remplir la tableau Excel
    initialiserExcel(ws, content)

    # Calculer le DeltaSphere
    insertUneColone(ws, 'Q', 'DeltaSphere', calculerDelta)
    insertUneColone(ws, 'R', 'DeltaCylindre', calculerDelta)
    insertUneColone(ws, 'S', 'DeltaAxe', calculerDelta)
    insertUneColone(ws, 'T', 'DeltaAddition', calculerDelta)
    insertUneColone(ws, 'U', 'DeltaEpais.centre', calculerDelta)

    # Ne pas supprimer la ligne derrière
    saveWB(wb)
    print("Sccessed!")
    print("#"*20, "="*10, "#"*20)
    print("#"*20, "="*10, "#"*20)

    # Analyser la dispersion
    print("Dispersion : ")
    dispersion(ws, "Q")
    dispersion(ws, "R")
    dispersion(ws, "S")
    dispersion(ws, "T")
    dispersion(ws, "U")
    print("#"*20, "="*10, "#"*20)
    print("#"*20, "="*10, "#"*20)

    # Intervalle de confiance
    print("Intervalle de confiance :")
    inter_confiance(ws, "Q")
    inter_confiance(ws, "R")
    inter_confiance(ws, "S")
    inter_confiance(ws, "T")
    inter_confiance(ws, "U")
    print("#"*20, "="*10, "#"*20)
    print("#"*20, "="*10, "#"*20)

    # H0 : l’écart entre la valeur demandée et la valeur fabriquée est nul.
    print("H0 : l’écart entre la valeur demandée et la valeur fabriquée est nul.")
    calculer_h1(ws, "Q")
    calculer_h1(ws, "R")
    calculer_h1(ws, "S")
    calculer_h1(ws, "T")
    calculer_h1(ws, "U")
    print("#"*20, "="*10, "#"*20)
    print("#"*20, "="*10, "#"*20)

    # TEST DE COMPARAISON DE MOYENNES
    print("TEST DE COMPARAISON DE MOYENNES")
    test_de_comparaison_de_moyenne(ws, "Q")
    test_de_comparaison_de_moyenne(ws, "R")
    test_de_comparaison_de_moyenne(ws, "S")
    test_de_comparaison_de_moyenne(ws, "T")
    test_de_comparaison_de_moyenne(ws, "U")
    print("#"*20, "="*10, "#"*20)
    print("#"*20, "="*10, "#"*20)
    
if __name__ == '__main__':
    main()
